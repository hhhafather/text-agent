import os
import uuid
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import streamlit as st
from langchain_community.document_loaders import PyPDFLoader, Docx2txtLoader, TextLoader
from utils import dataframe_agent

plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 页面配置
st.set_page_config(
    page_title="📊 智能文档分析助手",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://github.com',
        'Report a bug': "https://github.com",
        'About': "# 智能文档分析助手\n基于AI的文档数据分析工具"
    }
)



# 自定义CSS样式
st.markdown("""
<style>
    .main {
        padding-top: 2rem;
    }
    .stApp {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
    .block-container {
        padding-top: 1rem;
        padding-bottom: 0rem;
        padding-left: 5rem;
        padding-right: 5rem;
        background-color: rgba(255, 255, 255, 0.95);
        border-radius: 15px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        margin-top: 2rem;
    }
    .sidebar .sidebar-content {
        background: linear-gradient(180deg, #667eea 0%, #764ba2 100%);
        border-radius: 10px;
    }
    .stButton > button {
        background: linear-gradient(45deg, #667eea, #764ba2);
        color: white;
        border: none;
        border-radius: 25px;
        padding: 0.5rem 2rem;
        font-weight: bold;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6);
    }
    .stTextArea > div > div > textarea {
        border-radius: 10px;
        border: 2px solid #e0e0e0;
        transition: border-color 0.3s ease;
    }
    .stTextArea > div > div > textarea:focus {
        border-color: #667eea;
        box-shadow: 0 0 10px rgba(102, 126, 234, 0.3);
    }
    .stExpander {
        border: 1px solid #e0e0e0;
        border-radius: 10px;
        overflow: hidden;
    }
    .metric-card {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        margin: 0.5rem 0;
    }
    .upload-section {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
        padding: 2rem;
        border-radius: 15px;
        margin: 1rem 0;
        border: 2px dashed #667eea;
    }
    .title-container {
        text-align: center;
        padding: 2rem 0;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
</style>
""", unsafe_allow_html=True)

plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False


def create_chart(input_data, chart_type):
    """生成统计图表"""
    df_data = pd.DataFrame(
        data={
            "x": input_data["columns"],
            "y": input_data["data"]
        }
    ).set_index("x")
    
    if chart_type == "柱状图":
        plt.style.use('seaborn-v0_8')
        fig, ax = plt.subplots(figsize=(10, 6), dpi=120)
        bars = ax.bar(input_data["columns"], input_data["data"], 
                     color=['#667eea', '#764ba2', '#f093fb', '#f5576c', '#4facfe'],
                     alpha=0.8, edgecolor='white', linewidth=2)
        ax.set_title('数据分析结果', fontsize=16, fontweight='bold', pad=20)
        ax.grid(True, alpha=0.3)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        st.pyplot(fig)
        
    elif chart_type == "折线图":
        st.line_chart(df_data, use_container_width=True)
        
    elif chart_type == "饼图":
        plt.style.use('seaborn-v0_8')
        fig, ax = plt.subplots(figsize=(8, 8), dpi=120)
        colors = ['#667eea', '#764ba2', '#f093fb', '#f5576c', '#4facfe']
        wedges, texts, autotexts = ax.pie(input_data["data"], labels=input_data["columns"], 
                                         autopct='%1.1f%%', startangle=90, colors=colors,
                                         explode=[0.05]*len(input_data["data"]),
                                         shadow=True)
        ax.set_title('数据分布图', fontsize=16, fontweight='bold', pad=20)
        plt.tight_layout()
        st.pyplot(fig)


# 使用 st.cache_data 缓存文件加载函数
@st.cache_data(show_spinner="正在加载数据...")
def load_data(file_path, file_type, sheet_name=None):
    """根据文件类型加载数据并返回DataFrame"""
    try:
        if file_type == 'xlsx' or file_type == 'xls':
            return pd.read_excel(file_path, sheet_name=sheet_name)
        elif file_type == 'csv':
            return pd.read_csv(file_path)
        elif file_type == 'pdf':
            loader = PyPDFLoader(file_path)
            documents = loader.load()
            return pd.DataFrame({"Content": ["\n".join([doc.page_content for doc in documents])]})
        elif file_type == 'docx':
            loader = Docx2txtLoader(file_path)
            documents = loader.load()
            return pd.DataFrame({"Content": ["\n".join([doc.page_content for doc in documents])]})
        elif file_type == 'txt' or file_type == 'md':
            # 【修复】智能尝试多种编码加载文本文件
            try:
                # 优先尝试 UTF-8，因为它是最标准的编码
                loader = TextLoader(file_path, encoding='utf-8')
                documents = loader.load()
            except (UnicodeDecodeError, RuntimeError):
                # 如果 UTF-8 失败，回退尝试 GBK 编码，它在中国很常用
                loader = TextLoader(file_path, encoding='gbk')
                documents = loader.load()
            return pd.DataFrame({"Content": ["\n".join([doc.page_content for doc in documents])]})
        else:
            st.error(f"不支持的文件类型: {file_type}")
            return pd.DataFrame()
    except Exception as e:
        st.error(f"加载文件时发生错误: {e}")
        return pd.DataFrame()




# 主标题区域
st.markdown("""
<div class="title-container">
    <h1 style="font-size: 3rem; margin-bottom: 0;">📊 智能文档分析助手</h1>
    <p style="font-size: 1.2rem; opacity: 0.8; margin-top: 0.5rem;">基于AI的智能数据分析与可视化平台</p>
</div>
""", unsafe_allow_html=True)

# 功能介绍卡片
col1, col2, col3, col4 = st.columns(4)
with col1:
    st.markdown("""
    <div class="metric-card">
        <h3 style="color: #667eea; margin-bottom: 0.5rem;">📁 多格式支持</h3>
        <p style="font-size: 0.9rem; color: #666;">Excel, CSV, PDF, Word, TXT</p>
    </div>
    """, unsafe_allow_html=True)

with col2:
    st.markdown("""
    <div class="metric-card">
        <h3 style="color: #764ba2; margin-bottom: 0.5rem;">🤖 AI分析</h3>
        <p style="font-size: 0.9rem; color: #666;">智能数据洞察与解答</p>
    </div>
    """, unsafe_allow_html=True)

with col3:
    st.markdown("""
    <div class="metric-card">
        <h3 style="color: #f093fb; margin-bottom: 0.5rem;">📊 可视化</h3>
        <p style="font-size: 0.9rem; color: #666;">柱状图、折线图、饼图</p>
    </div>
    """, unsafe_allow_html=True)

with col4:
    st.markdown("""
    <div class="metric-card">
        <h3 style="color: #f5576c; margin-bottom: 0.5rem;">⚡ 实时处理</h3>
        <p style="font-size: 0.9rem; color: #666;">快速响应与分析</p>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

if 'session_id' not in st.session_state:
    st.session_state['session_id'] = uuid.uuid4().hex
    st.session_state['is_new_file'] = True
    st.session_state['current_file_name'] = None

# 如果 'uploads' 目录不存在则创建它
if not os.path.exists('uploads'):
    os.makedirs('uploads')

with st.sidebar:
    st.markdown("""
    <div style="text-align: center; padding: 1rem; margin-bottom: 2rem;">
        <h2 style="color: white; margin-bottom: 0.5rem;">🚀 开始分析</h2>
        <p style="color: rgba(255,255,255,0.8); font-size: 0.9rem;">选择文件类型并上传数据</p>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("### 📂 文件类型选择")
    option = st.radio(
        "请选择数据文件类型:", 
        ("Excel", "CSV", "txt", "pdf", "docx", "md"),
        help="支持多种常见文件格式的智能解析"
    )

    file_type_map = {
        "Excel": ["xlsx", "xls"],
        "CSV": ["csv"],
        "txt": ["txt"],
        "pdf": ["pdf"],
        "docx": ["docx"],
        "md": ["md"]
    }

    # 根据所选选项确定文件上传器允许的文件类型
    allowed_file_types = file_type_map.get(option, ["csv"])
    
    st.markdown("### 📤 文件上传")
    data = st.file_uploader(
        f"📁 上传你的{option}数据文件", 
        type=allowed_file_types,
        help=f"支持的格式: {', '.join(allowed_file_types)}",
        accept_multiple_files=False
    )
    
    if data:
        st.success(f"✅ 文件已上传: {data.name}")
        file_size = len(data.getvalue()) / 1024  # KB
        st.info(f"📊 文件大小: {file_size:.1f} KB")
    
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; padding: 1rem; color: rgba(255,255,255,0.7);">
        <small>💡 提示: 上传文件后，在右侧输入您的问题</small>
    </div>
    """, unsafe_allow_html=True)

# 清除缓存的 DataFrame 如果上传了新文件
if data:
    if 'current_file_name' not in st.session_state or st.session_state['current_file_name'] != data.name:
        st.session_state['is_new_file'] = True
        st.session_state['current_file_name'] = data.name

        suffix = data.name[data.name.rfind('.'):].lower().replace('.', '')

        temp_file_path = os.path.join('uploads', f'{st.session_state["session_id"]}_{data.name}')

        # 【修复】统一使用二进制写入模式('wb')保存所有上传文件，以保留其原始字节，避免编码问题
        with open(temp_file_path, 'wb') as temp_file:
            temp_file.write(data.read())

        sheet_name_to_load = None
        if suffix in ('xlsx', 'xls'):
            try:
                wb = openpyxl.load_workbook(temp_file_path)
                sheet_names = wb.sheetnames
                if sheet_names:
                    if 'selected_excel_sheet' in st.session_state and st.session_state[
                        'selected_excel_sheet'] in sheet_names:
                        default_sheet_index = sheet_names.index(st.session_state['selected_excel_sheet'])
                    else:
                        default_sheet_index = 0

                    selected_sheet = st.radio(label="请选择要加载的工作表：", options=sheet_names,
                                              index=default_sheet_index, key="excel_sheet_selector")
                    st.session_state['selected_excel_sheet'] = selected_sheet
                    sheet_name_to_load = selected_sheet
                else:
                    st.warning("Excel 文件中没有检测到工作表。")
            except Exception as e:
                st.error(f"读取Excel工作表时出错: {e}")

        st.session_state["df"] = load_data(temp_file_path, suffix, sheet_name=sheet_name_to_load)

    elif 'current_file_name' in st.session_state and st.session_state['current_file_name'] == data.name:
        suffix = data.name[data.name.rfind('.'):].lower().replace('.', '')
        if suffix in ('xlsx', 'xls'):
            temp_file_path = os.path.join('uploads', f'{st.session_state["session_id"]}_{data.name}')
            try:
                wb = openpyxl.load_workbook(temp_file_path)
                sheet_names = wb.sheetnames
                if sheet_names:
                    if 'selected_excel_sheet' in st.session_state and st.session_state[
                        'selected_excel_sheet'] in sheet_names:
                        default_sheet_index = sheet_names.index(st.session_state['selected_excel_sheet'])
                    else:
                        default_sheet_index = 0
                    selected_sheet = st.radio(label="请选择要加载的工作表：", options=sheet_names,
                                              index=default_sheet_index, key="excel_sheet_selector_re_render")
                    if selected_sheet != st.session_state.get('selected_excel_sheet'):
                        st.session_state['selected_excel_sheet'] = selected_sheet
                        st.session_state["df"] = load_data(temp_file_path, suffix, sheet_name=selected_sheet)
                else:
                    st.warning("Excel 文件中没有检测到工作表。")
            except Exception as e:
                st.error(f"读取Excel工作表时出错: {e}")

# 数据展示区域
if "df" in st.session_state and not st.session_state['df'].empty:
    st.markdown("## 📋 数据预览")
    
    # 数据统计信息
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("📊 总行数", len(st.session_state["df"]), help="数据集的总行数")
    with col2:
        st.metric("📈 总列数", len(st.session_state["df"].columns), help="数据集的总列数")
    with col3:
        memory_usage = st.session_state["df"].memory_usage(deep=True).sum() / 1024
        st.metric("💾 内存占用", f"{memory_usage:.1f} KB", help="数据在内存中的占用大小")
    with col4:
        null_count = st.session_state["df"].isnull().sum().sum()
        st.metric("❓ 缺失值", null_count, help="数据集中的缺失值总数")
    
    with st.expander("🔍 查看原始数据", expanded=False):
        st.dataframe(
            st.session_state["df"], 
            use_container_width=True,
            height=400
        )
        
        # 数据类型信息
        if st.checkbox("显示数据类型信息"):
            st.subheader("📝 数据类型")
            dtype_df = pd.DataFrame({
                '列名': st.session_state["df"].columns,
                '数据类型': st.session_state["df"].dtypes.values,
                '非空值数量': st.session_state["df"].count().values
            })
            st.dataframe(dtype_df, use_container_width=True)
            
elif "df" in st.session_state and "current_file_name" in st.session_state and st.session_state[
    'current_file_name'] is not None and st.session_state['df'].empty:
    st.warning("⚠️ 上传的文件已处理，但生成的 DataFrame 为空。")
elif data is None and "df" in st.session_state:
    st.session_state.pop('df', None)

# 查询输入区域
st.markdown("## 💬 智能问答")
st.markdown("""
<div style="background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%); 
           padding: 1.5rem; border-radius: 15px; margin: 1rem 0; 
           border-left: 4px solid #667eea;">
    <h4 style="margin-top: 0; color: #667eea;">🤖 AI助手使用指南</h4>
    <ul style="margin-bottom: 0; color: #555;">
        <li>📊 <strong>数据分析</strong>: "分析销售数据的趋势"</li>
        <li>📈 <strong>可视化</strong>: "生成销售额的柱状图"</li>
        <li>🔍 <strong>数据查询</strong>: "找出销售额最高的产品"</li>
        <li>📋 <strong>统计信息</strong>: "计算平均值和总和"</li>
    </ul>
</div>
""", unsafe_allow_html=True)

query = st.text_area(
    "💭 请输入你关于数据的问题或可视化需求：",
    disabled="df" not in st.session_state,
    placeholder='例如：分析销售数据的趋势，或生成销售额的柱状图...',
    height=100,
    help="输入您想要了解的数据问题，AI将为您提供详细分析"
)

col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    button = st.button(
        "🚀 开始分析", 
        use_container_width=True,
        help="点击开始AI分析",
        disabled=not query or "df" not in st.session_state
    )

if button and not data:
    st.error("❌ 请先上传数据文件")
    st.stop()

if button and query and "df" in st.session_state and not st.session_state["df"].empty:
    # 显示分析进度
    with st.spinner("🤖 AI正在分析您的数据，请稍候..."):
        @st.cache_data(show_spinner=False, ttl=3600)
        def get_analysis_result(df_hash, query_text):
            return dataframe_agent(st.session_state["df"], query_text)

        df_hash_key = hash((st.session_state["df"].shape, tuple(st.session_state["df"].columns)))
        result = get_analysis_result(df_hash_key, query)
    
    # 结果展示区域
    st.markdown("## 🎯 分析结果")
    
    if "answer" in result:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #e8f5e8 0%, #f0f8f0 100%); 
                   padding: 1.5rem; border-radius: 15px; margin: 1rem 0;
                   border-left: 4px solid #4CAF50;">
            <h4 style="margin-top: 0; color: #2E7D32;">💡 AI分析结果</h4>
        </div>
        """, unsafe_allow_html=True)
        st.markdown(f"**{result['answer']}**")
        
    if "table" in result:
        st.markdown("### 📊 数据表格")
        result_df = pd.DataFrame(result["table"]["data"], columns=result["table"]["columns"])
        st.dataframe(result_df, use_container_width=True)
        
        # 添加下载按钮
        csv = result_df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="📥 下载表格数据",
            data=csv,
            file_name='analysis_result.csv',
            mime='text/csv'
        )

    # 图表展示
    if "bar" in result:
        st.markdown("### 📊 柱状图分析")
        create_chart(result["bar"], "柱状图")
    elif "line" in result:
        st.markdown("### 📈 趋势分析")
        create_chart(result["line"], "折线图")
    elif "pie" in result:
        st.markdown("### 🥧 分布分析")
        create_chart(result["pie"], "饼图")
        
    # 成功提示
    st.success("✅ 分析完成！如需进一步分析，请输入新的问题。")
    
elif button:
    st.warning("⚠️ 请上传数据文件并输入问题。")

# 页脚
st.markdown("---")
st.markdown("""
<div style="text-align: center; padding: 2rem; color: #666;">
    <p style="margin: 0;">🚀 <strong>智能文档分析助手</strong> | 让数据分析变得简单高效</p>
    <p style="margin: 0.5rem 0 0 0; font-size: 0.9rem;">Powered by AI • Built with ❤️</p>
</div>
""", unsafe_allow_html=True)
